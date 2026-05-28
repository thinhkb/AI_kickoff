"""Deep comparison of old vs new scoring for regression analysis."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

# New regressions: 701, 716, 718, 750
# Fixed: 726, 729
# Still broken: 707, 708, 709, 715, 719, 725, 731, 733, 736, 739, 741, 742, 743
for mid in [701, 716, 718, 750, 715, 733, 742]:
    mid_float = float(mid)
    if mid_float not in questions:
        continue
    q_data = questions[mid_float]
    question = q_data["fun_question"]
    note = q_data.get("note", "") or ""
    
    options = parse_options(note)
    if not options:
        continue
    
    is_neg = OptionScorer.is_negation_question(question)
    
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
    chunks = [chunk for chunk, score in results]
    
    scores = scorer.score_options(
        question=question,
        options=options,
        evidence_chunks=chunks,
        retriever=retriever,
    )
    
    # Check negation logic
    if is_neg:
        sorted_asc = sorted(scores.items(), key=lambda x: (x[1], x[0]))
        selected = sorted_asc[0][0]
    else:
        sorted_desc = sorted(scores.items(), key=lambda x: (-x[1], x[0]))
        selected = sorted_desc[0][0]
    
    print(f"\nID: {mid} | negation={is_neg} | selected={selected}")
    print(f"Q: {question[:100]}")
    for letter in sorted(options.keys()):
        print(f"  {letter}: {options[letter][:80]} | score={scores.get(letter, 0):.4f}")
    print(f"  Top evidence docs: {[c.doc_id for c, _ in results[:3]]}")
