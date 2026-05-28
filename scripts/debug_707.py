import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

# ID 707
q_data = questions[707.0]
question = q_data["fun_question"]
note = q_data["note"]
options = parse_options(note)

results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
chunks = [chunk for chunk, score in results]

print("=== retrieved top 5 ===")
for i, c in enumerate(chunks[:5]):
    print(f"[{i}] {c.doc_id} {c.page}")
    print(c.text[:200])
    print()

overlap_scores = scorer._score_with_normalized_overlap(question, options, chunks[:20])
tfidf_scores = scorer._score_with_tfidf_sklearn(question, options, chunks[:25])

overlap_norm = scorer._minmax(overlap_scores)
tfidf_norm = scorer._minmax(tfidf_scores)

print("Overlap scores:", overlap_scores)
print("Overlap norm:", overlap_norm)
print("TFIDF scores:", tfidf_scores)
print("TFIDF norm:", tfidf_norm)

qopt_scores = {}
opt_scores = {}
doc_hint = scorer._extract_doc_hint(question)
for letter, option_text in options.items():
    qopt_query = f"{question} {option_text}"
    qopt_results = retriever.retrieve_with_metadata_filter(qopt_query, top_k=5)
    qopt_scores[letter] = scorer._ranked_score(qopt_results)

    opt_query = f"{doc_hint} {option_text}".strip() if doc_hint else option_text
    opt_results = retriever.retrieve_with_metadata_filter(opt_query, top_k=5)
    opt_scores[letter] = scorer._ranked_score(opt_results)

print("qopt scores:", qopt_scores)
print("qopt norm:", scorer._minmax(qopt_scores))
print("opt scores:", opt_scores)
print("opt norm:", scorer._minmax(opt_scores))
