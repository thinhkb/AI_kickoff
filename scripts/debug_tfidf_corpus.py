import sys, json, re
from pathlib import Path
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity as sk_cosine
import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Load chunks
chunks = []
with open(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl", "r", encoding="utf-8") as f:
    for line in f:
        chunks.append(json.loads(line.strip()))

texts = [c["text"] for c in chunks]

print("Fitting TfidfVectorizer on corpus...")
vectorizer = TfidfVectorizer(
    analyzer='word',
    token_pattern=r'\w{2,}',
    sublinear_tf=True,
    norm=None,
)
vectorizer.fit(texts)
print("Vocabulary size:", len(vectorizer.vocabulary_))

# Let's see some top words by IDF
idfs = vectorizer.idf_
vocab = vectorizer.vocabulary_
words_by_idf = sorted(vocab.keys(), key=lambda w: idfs[vocab[w]])
print("Most common words (lowest IDF):", words_by_idf[:20])
print("Rarest words (highest IDF):", words_by_idf[-20:])

# Now let's test scoring ID 707
from src.document.option_parser import parse_options
import openpyxl

wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

q_data = questions[707.0]
question = q_data["fun_question"]
note = q_data["note"]
options = parse_options(note)

# Retrieve top 25 chunks
from src.document.retriever import DocumentRetriever
retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
chunks_ret = [chunk for chunk, score in results]
evidence_combined = " ".join(c.text for c in chunks_ret[:25])

evidence_vec = vectorizer.transform([evidence_combined])
evidence_norm = float(np.linalg.norm(evidence_vec.toarray()))

print("\n=== Corpus-based TF-IDF Cosine Similarity ===")
for letter, text in sorted(options.items()):
    opt_vec = vectorizer.transform([text])
    opt_norm = float(np.linalg.norm(opt_vec.toarray()))
    sim = float(sk_cosine(opt_vec, evidence_vec)[0][0])
    dot_prod = sim * opt_norm * evidence_norm
    soft_sim = dot_prod / (opt_norm ** 0.5 + 1e-9)
    print(f"  {letter}: {text[:50]} -> Sim: {sim:.4f} | Norm: {opt_norm:.4f} | DotProd: {dot_prod:.4f} | SoftSim: {soft_sim:.4f}")
