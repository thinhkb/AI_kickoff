"""Deep-dive into scoring behavior for mismatched questions."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

# Load retriever
retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

# Load example data
import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

ws_res = wb["example_result"]
headers_res = [c.value for c in ws_res[1]]
gt_answers = {}
for row in ws_res.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_res, row))
    if d.get("id") is not None:
        gt_answers[float(d["id"])] = d["func_param"]

# Clear output file
with open(PROJECT_ROOT / "outputs" / "debug_scoring_out.txt", "w", encoding="utf-8") as f:
    f.write("DEBUG SCORING RESULTS\n")

# Mismatched IDs
mismatch_ids = []
try:
    with open(PROJECT_ROOT / "outputs" / "eval_mismatches.jsonl", "r", encoding="utf-8") as f:
        for line in f:
            d = json.loads(line.strip())
            if d.get("issue") == "document_answer":
                mismatch_ids.append(d["id"])
except Exception as e:
    mismatch_ids = [707, 708, 709, 715, 719, 725, 726, 729, 731, 733, 736, 739, 741, 742, 743]

for mid in mismatch_ids:
    mid_float = float(mid)
    if mid_float not in questions:
        continue
    q_data = questions[mid_float]
    question = q_data["fun_question"]
    note = q_data.get("note", "") or ""
    gt_answer = gt_answers.get(mid_float, "")
    
    out_lines = []
    out_lines.append(f"\n{'='*80}")
    out_lines.append(f"ID: {mid} | Q: {question[:100]}")
    
    # Parse options
    options = parse_options(note)
    if not options:
        out_lines.append("  NO OPTIONS PARSED!")
        with open(PROJECT_ROOT / "outputs" / "debug_scoring_out.txt", "a", encoding="utf-8") as out_f:
            out_f.write("\n".join(out_lines) + "\n")
        continue
    
    # Retrieve chunks
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
    chunks = [chunk for chunk, score in results]
    
    # Show top retrieved chunks
    out_lines.append(f"  Retrieved {len(chunks)} chunks")
    for i, (chunk, score) in enumerate(results[:3]):
        out_lines.append(f"  Top-{i+1} (score={score:.4f}): doc={chunk.doc_id} page={chunk.page}")
        out_lines.append(f"    text={chunk.text[:150]}...")
    
    # Score options
    scores = scorer.score_options(
        question=question,
        options=options,
        evidence_chunks=chunks,
        retriever=retriever,
    )
    
    out_lines.append(f"  Scores: {scores}")
    out_lines.append(f"  Options:")
    for letter, text in sorted(options.items()):
        marker = " <<< CORRECT" if letter == gt_answer else ""
        out_lines.append(f"    {letter}: {text[:100]}{marker}")

    with open(PROJECT_ROOT / "outputs" / "debug_scoring_out.txt", "a", encoding="utf-8") as out_f:
        out_f.write("\n".join(out_lines) + "\n")
