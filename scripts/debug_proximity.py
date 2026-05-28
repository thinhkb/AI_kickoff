import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

def custom_tokenize(text: str) -> list[str]:
    stopwords = {
        "và", "là", "của", "có", "trong", "theo", "được", "từ",
        "một", "những", "các", "về", "để", "nào", "gì", "với",
        "khi", "nếu", "thì", "cho", "trên", "dưới", "bao", "nhiêu",
        "đâu", "hãy", "hay", "không", "phải", "chính", "chủ", "yếu",
        "đã", "bị", "vào", "ra", "ở", "này", "đó",
    }
    return [
        t for t in re.findall(r"\w+", (text or "").lower(), flags=re.UNICODE)
        if (len(t) > 1 or t.isdigit()) and t not in stopwords
    ]

def score_with_proximity(question, options, evidence_chunks, mid=0.0):
    evidence_text = " ".join(c.text for c in evidence_chunks)
    sentences = re.split(r"[.!?\n]+", evidence_text)
    
    q_tokens = custom_tokenize(question)
    q_keywords = [t for t in q_tokens if len(t) > 1 or t.isdigit()]
    if not q_keywords:
        return {k: 0.0 for k in options}

    scores = {}
    for letter, option_text in options.items():
        opt_tokens = custom_tokenize(option_text)
        useful_tokens = [t for t in opt_tokens if t not in q_tokens]
        if not useful_tokens:
            useful_tokens = opt_tokens
        
        max_score = 0.0
        for sent in sentences:
            sent_words = re.findall(r"\w+", sent.lower(), flags=re.UNICODE)
            if not sent_words:
                continue
            
            opt_indices = [i for i, w in enumerate(sent_words) if w in useful_tokens]
            if not opt_indices:
                continue
            
            q_indices = {kw: [i for i, w in enumerate(sent_words) if w == kw] for kw in q_keywords}
            
            sent_score = 0.0
            for kw, indices in q_indices.items():
                if not indices:
                    continue
                min_dist = min(abs(q_idx - opt_idx) for q_idx in indices for opt_idx in opt_indices)
                sent_score += 1.0 / (min_dist + 1.0)
            
            if sent_score > max_score:
                if mid == 716.0:
                    print(f"    [716 Match] sent='{sent[:60]}...' | opt_indices={opt_indices} | sent_words={sent_words} | sent_score={sent_score:.4f} | useful_tokens={useful_tokens}")
                max_score = sent_score
        
        scores[letter] = max_score
    return scores

# Test on tricky IDs
for mid in [708.0, 716.0, 725.0, 726.0, 728.0, 730.0, 731.0]:
    q_data = questions[mid]
    question = q_data["fun_question"]
    note = q_data["note"] or ""
    options = parse_options(note)
    
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=25)
    chunks = [chunk for chunk, score in results]
    
    scores = score_with_proximity(question, options, chunks, mid=mid)
    print(f"\n=== Proximity Scores for {mid} ===")
    print(f"Question: {question[:80]}...")
    for k, v in sorted(scores.items()):
        print(f"  {k}: {options.get(k, '')[:50]} -> Score: {v:.4f}")
