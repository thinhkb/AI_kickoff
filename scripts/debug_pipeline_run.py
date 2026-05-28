import sys, json
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.pipeline import Pipeline
from src.schemas import InputSample
import openpyxl

wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

pipeline = Pipeline()
pipeline.load()

# Test ID 708
mid = 708.0
q_data = questions[mid]
sample = InputSample.from_dict(q_data)

print("Running pipeline.predict_one for ID 708...")
pred = pipeline.predict_one(sample)
print("Prediction Result:", pred.to_dict())

# Test ID 701
mid = 701.0
q_data = questions[mid]
sample = InputSample.from_dict(q_data)

print("\nRunning pipeline.predict_one for ID 701...")
pred = pipeline.predict_one(sample)
print("Prediction Result:", pred.to_dict())
