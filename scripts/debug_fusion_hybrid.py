import sys, json, re
from pathlib import Path
from sklearn.feature_extraction.text import TfidfVectorizer
import numpy as np

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

# Load retriever
retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

# Fit corpus TF-IDF
chunks = retriever.chunks
texts = [c.text for c in chunks]
vectorizer = TfidfVectorizer(
    analyzer='word',
    token_pattern=r'\w{2,}',
    sublinear_tf=True,
    norm=None,
)
vectorizer.fit(texts)

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    if d.get("id") is not None:
        questions[float(d["id"])] = d

ws_res = wb["example_result"]
headers_res = [c.value for c in ws_res[1]]
gt_answers = {}
for row in ws_res.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers_res, row))
    if d.get("id") is not None:
        gt_answers[float(d["id"])] = d["func_param"]

def custom_tokenize(text: str) -> list[str]:
    stopwords = {
        "và", "là", "của", "có", "trong", "theo", "được", "từ",
        "một", "những", "các", "về", "để", "nào", "gì", "với",
        "khi", "nếu", "thì", "cho", "trên", "dưới", "bao", "nhiêu",
        "đâu", "hãy", "hay", "không", "phải", "chính", "chủ", "yếu",
        "đã", "bị", "vào", "ra", "ở", "này", "đó",
    }
    return [
        t for t in re.findall(r"\w+", (text or "").lower(), flags=re.UNICODE)
        if (len(t) > 1 or t.isdigit()) and t not in stopwords
    ]

def score_with_proximity(question, options, evidence_chunks):
    evidence_text = " ".join(c.text for c in evidence_chunks)
    sentences = re.split(r"[.!?\n]+", evidence_text)
    
    q_tokens = custom_tokenize(question)
    q_keywords = [t for t in q_tokens if len(t) > 1 or t.isdigit()]
    if not q_keywords:
        return {k: 0.0 for k in options}

    scores = {}
    for letter, option_text in options.items():
        opt_tokens = custom_tokenize(option_text)
        useful_tokens = [t for t in opt_tokens if t not in q_tokens]
        if not useful_tokens:
            useful_tokens = opt_tokens
        
        max_score = 0.0
        for sent in sentences:
            sent_words = re.findall(r"\w+", sent.lower(), flags=re.UNICODE)
            if not sent_words:
                continue
            
            opt_indices = [i for i, w in enumerate(sent_words) if w in useful_tokens]
            if not opt_indices:
                continue
            
            q_indices = {kw: [i for i, w in enumerate(sent_words) if w == kw] for kw in q_keywords}
            
            sent_score = 0.0
            for kw, indices in q_indices.items():
                if not indices:
                    continue
                min_dist = min(abs(q_idx - opt_idx) for q_idx in indices for opt_idx in opt_indices)
                sent_score += 1.0 / (min_dist + 1.0)
            
            if sent_score > max_score:
                max_score = sent_score
        
        scores[letter] = max_score
    return scores

def score_with_normalized_overlap(question, options, evidence_chunks):
    evidence = " ".join(c.text for c in evidence_chunks)
    evidence_tokens = set(custom_tokenize(evidence))
    question_tokens = set(custom_tokenize(question))
    evidence_lower = evidence.lower()

    scores = {}
    for letter, option_text in options.items():
        option_tokens = set(custom_tokenize(option_text))
        useful_tokens = option_tokens - question_tokens
        if not useful_tokens:
            useful_tokens = option_tokens
        
        # Proportional length-normalized overlap (using square root to avoid short option bias)
        overlap = len(useful_tokens & evidence_tokens) / max(len(useful_tokens) ** 0.5, 1.0)
        phrase_bonus = 1.0 if option_text.lower() in evidence_lower else 0.0
        scores[letter] = overlap + phrase_bonus
    return scores

def minmax(scores):
    if not scores:
        return {}
    values = list(scores.values())
    min_v, max_v = min(values), max(values)
    span = max_v - min_v
    if span <= 1e-12:
        return {k: 0.0 for k in scores}
    return {k: (v - min_v) / span for k, v in scores.items()}

# Tricky IDs
tricky_ids = [707.0, 708.0, 709.0, 715.0, 716.0, 725.0, 726.0, 728.0, 730.0, 731.0, 733.0, 736.0, 739.0, 741.0, 743.0]

for mid in tricky_ids:
    q_data = questions[mid]
    question = q_data["fun_question"]
    note = q_data["note"] or ""
    options = parse_options(note)
    gt_answer_raw = gt_answers.get(mid, "")
    gt_answer = gt_answer_raw
    try:
        gt_answer = json.loads(gt_answer_raw).get("result", gt_answer_raw)
    except:
        pass
    
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
    chunks = [chunk for chunk, score in results]
    
    overlap_scores = score_with_normalized_overlap(question, options, chunks[:20])
    proximity_scores = score_with_proximity(question, options, chunks[:25])
    
    evidence_combined = " ".join(c.text for c in chunks[:25])
    evidence_vec = vectorizer.transform([evidence_combined])
    tfidf_scores = {}
    for letter, text in sorted(options.items()):
        opt_vec = vectorizer.transform([text])
        tfidf_scores[letter] = float(opt_vec.dot(evidence_vec.T).toarray()[0][0])
        
    qopt_scores = {}
    opt_scores = {}
    doc_hint = scorer._extract_doc_hint(question)
    for letter, option_text in options.items():
        qopt_query = f"{question} {option_text}"
        qopt_results = retriever.retrieve_with_metadata_filter(qopt_query, top_k=5)
        qopt_scores[letter] = scorer._ranked_score(qopt_results)

        opt_query = f"{doc_hint} {option_text}".strip() if doc_hint else option_text
        opt_results = retriever.retrieve_with_metadata_filter(opt_query, top_k=5)
        opt_scores[letter] = scorer._ranked_score(opt_results)
        
    overlap_norm = minmax(overlap_scores)
    proximity_norm = minmax(proximity_scores)
    tfidf_norm = minmax(tfidf_scores)
    qopt_norm = minmax(qopt_scores)
    opt_norm = minmax(opt_scores)
    
    # 5-way hybrid fusion with dynamic weights for short/numeric options
    is_short_options = all(len(re.findall(r"\w+", opt.lower(), flags=re.UNICODE)) <= 2 for opt in options.values())
    
    scores = {}
    if is_short_options:
        for letter in options:
            scores[letter] = (
                0.05 * overlap_norm.get(letter, 0.0)
                + 0.85 * proximity_norm.get(letter, 0.0)
                + 0.05 * tfidf_norm.get(letter, 0.0)
                + 0.05 * qopt_norm.get(letter, 0.0)
            )
    else:
        for letter in options:
            scores[letter] = (
                0.25 * overlap_norm.get(letter, 0.0)
                + 0.20 * proximity_norm.get(letter, 0.0)
                + 0.15 * tfidf_norm.get(letter, 0.0)
                + 0.25 * qopt_norm.get(letter, 0.0)
                + 0.15 * opt_norm.get(letter, 0.0)
            )
        
    is_neg = scorer.is_negation_question(question)
    best = scorer.select_best(scores, is_negation=is_neg)
    
    correct = (best == gt_answer)
    marker = "✓ CORRECT" if correct else "✗ WRONG"
    print(f"\nID: {mid} | {marker} | Pred: {best} | GT: {gt_answer}")
    print(f"  Q: {question[:80]}...")
    print(f"  Scores: { {k: f'{v:.3f}' for k, v in sorted(scores.items())} }")
