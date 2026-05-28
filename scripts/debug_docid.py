"""Check which mismatched questions have doc_id in note but not in question."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

mismatched_ids = [707, 708, 709, 715, 719, 725, 726, 729, 731, 733, 736, 739, 741, 742, 743]

for mid in mismatched_ids:
    q_data = questions[float(mid)]
    question = q_data["fun_question"]
    note = q_data.get("note", "") or ""
    
    doc_id_q = retriever._detect_doc_id(question)
    doc_id_n = retriever._detect_doc_id(note)
    
    # Check if note contains doc ref but question doesn't
    has_doc_in_note_only = doc_id_n and not doc_id_q
    
    # Check retrieval quality
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=5)
    top_docs = set(c.doc_id for c, s in results[:5])
    
    print(f"ID {mid}: doc_q={doc_id_q} doc_n={doc_id_n} note_only={has_doc_in_note_only} top_docs={top_docs}")
    if has_doc_in_note_only:
        # Try retrieval with note doc_id
        results_n = retriever.retrieve_with_metadata_filter(query=question, doc_id=doc_id_n, top_k=5)
        top_docs_n = set(c.doc_id for c, s in results_n[:5])
        print(f"  -> With note doc_id: top_docs={top_docs_n}")
