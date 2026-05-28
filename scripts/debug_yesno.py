"""Test Yes/No handling and other specific fixes."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer
from src.document.document_solver import DocumentSolver

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()
solver = DocumentSolver(retriever=retriever, option_scorer=scorer)

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

# 739: Yes/No question
q_data = questions[739.0]
result = solver.solve(q_data["fun_question"], q_data.get("note", "") or "")
print(f"ID 739: result={result}")
print(f"  question: {q_data['fun_question'][:100]}")
print(f"  note: {q_data.get('note', '')}")
print()

# Check all mismatched IDs with full solve
for mid in [707, 708, 709, 713, 715, 716, 719, 725, 729, 731, 733, 736, 739, 741, 742, 743, 750]:
    mid_float = float(mid)
    if mid_float not in questions:
        continue
    q_data = questions[mid_float]
    result = solver.solve(q_data["fun_question"], q_data.get("note", "") or "")
    parsed = json.loads(result)
    print(f"ID {mid}: predicted={parsed['result']}")
