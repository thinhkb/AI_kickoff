"""Specifically debug id 713 regression."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

mid = 713
q_data = questions[float(mid)]
question = q_data["fun_question"]
note = q_data.get("note", "") or ""
options = parse_options(note)

print(f"Q: {question}")
print(f"Note: {note[:200]}")
print()

# Test with and without doc_id from note
doc_id_from_q = retriever._detect_doc_id(question)
doc_id_from_note = retriever._detect_doc_id(note)
print(f"doc_id from question: {doc_id_from_q}")
print(f"doc_id from note: {doc_id_from_note}")

# Retrieve with question doc_id only
results_q = retriever.retrieve_with_metadata_filter(query=question, doc_id=doc_id_from_q, top_k=10)
print(f"\nRetrieval with question doc_id ({doc_id_from_q}): {len(results_q)} chunks")
for i, (c, s) in enumerate(results_q[:3]):
    print(f"  [{i}] doc={c.doc_id} score={s:.2f} text={c.text[:100]}")

# Retrieve with note doc_id
if doc_id_from_note and doc_id_from_note != doc_id_from_q:
    results_n = retriever.retrieve_with_metadata_filter(query=question, doc_id=doc_id_from_note, top_k=10)
    print(f"\nRetrieval with note doc_id ({doc_id_from_note}): {len(results_n)} chunks")
    for i, (c, s) in enumerate(results_n[:3]):
        print(f"  [{i}] doc={c.doc_id} score={s:.2f} text={c.text[:100]}")

# Score with question-only retrieval
scores_q = scorer.score_options(question=question, options=options, evidence_chunks=[c for c,_ in results_q], retriever=retriever)
print(f"\nScores (question doc_id): {scores_q}")

# Score with note retrieval
if doc_id_from_note and doc_id_from_note != doc_id_from_q:
    scores_n = scorer.score_options(question=question, options=options, evidence_chunks=[c for c,_ in results_n], retriever=retriever)
    print(f"Scores (note doc_id): {scores_n}")
