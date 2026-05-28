"""Quick weight sweep to find best fusion weights for TF-IDF scoring."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

# Focus on the questions that changed between baseline and TF-IDF version
# Fixed: 713, 729, 743
# Broken: 702
# Check scores for these to understand
scorer = OptionScorer()

for mid in [702, 713, 729, 743]:
    mid_float = float(mid)
    if mid_float not in questions:
        continue
    q_data = questions[mid_float]
    question = q_data["fun_question"]
    note = q_data.get("note", "") or ""
    options = parse_options(note)
    if not options:
        continue
    
    results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
    chunks = [chunk for chunk, score in results]
    
    # Get individual component scores
    overlap_scores = scorer._score_with_normalized_overlap(question, options, chunks[:10])
    tfidf_scores = scorer._score_with_tfidf_sklearn(question, options, chunks[:15])
    
    doc_hint = scorer._extract_doc_hint(question)
    qopt_scores = {}
    opt_scores = {}
    for letter, option_text in options.items():
        qopt_query = f"{question} {option_text}"
        qopt_results = retriever.retrieve_with_metadata_filter(qopt_query, top_k=5)
        qopt_scores[letter] = scorer._ranked_score(qopt_results)
        opt_query = f"{doc_hint} {option_text}".strip() if doc_hint else option_text
        opt_results = retriever.retrieve_with_metadata_filter(opt_query, top_k=5)
        opt_scores[letter] = scorer._ranked_score(opt_results)
    
    overlap_norm = scorer._minmax(overlap_scores)
    tfidf_norm = scorer._minmax(tfidf_scores)
    qopt_norm = scorer._minmax(qopt_scores)
    opt_norm = scorer._minmax(opt_scores)
    
    print(f"ID {mid}:")
    for letter in sorted(options.keys()):
        print(f"  {letter}: overlap={overlap_norm.get(letter,0):.3f} tfidf={tfidf_norm.get(letter,0):.3f} qopt={qopt_norm.get(letter,0):.3f} opt={opt_norm.get(letter,0):.3f}")
    print(f"  Options: {[(k, v[:50]) for k,v in sorted(options.items())]}")
    print()
