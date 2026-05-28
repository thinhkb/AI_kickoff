"""
Deep analysis of evidence content for mismatched questions.
"""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

# Check specific: 715 (negation question about Public_049)
# GT=C (Thuật toán phải được thiết kế đặc thù, chuyên sâu thay vì tổng quát)
# This is the wrong statement - algorithms should be general, not specific

mid = 715
q_data = questions[float(mid)]
question = q_data["fun_question"]
note = q_data.get("note", "") or ""
options = parse_options(note)

results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
evidence = " ".join(c.text for c, _ in results[:10]).lower()

print("="*80)
print(f"ID: {mid}")
print(f"Q: {question}")
for letter, text in sorted(options.items()):
    opt_lower = text.lower()
    # Search for exact option text in evidence
    found = opt_lower in evidence
    # Also search for key phrases
    key_words = [w for w in re.findall(r'\w+', opt_lower) if len(w) > 2]
    matches = sum(1 for w in key_words if w in evidence)
    print(f"  {letter}: {text}")
    print(f"     exact_match={found}, keyword_matches={matches}/{len(key_words)}")

# Also look at what evidence says about "tổng quát" vs "đặc thù"
for term in ["tổng quát", "đặc thù", "rõ ràng", "hữu hạn", "dễ hiểu"]:
    count = evidence.count(term)
    print(f"  Evidence mentions '{term}': {count} times")

print()
print("TOP 5 EVIDENCE:")
for i, (chunk, score) in enumerate(results[:5]):
    print(f"  [{i}] score={score:.2f} doc={chunk.doc_id}")
    print(f"      {chunk.text[:200]}")
    print()

# Check id=733 (negation, Public_284 not apply)
mid = 733
q_data = questions[float(mid)]
question = q_data["fun_question"]
note = q_data.get("note", "") or ""
options = parse_options(note)

results = retriever.retrieve_with_metadata_filter(query=question, top_k=50)
evidence = " ".join(c.text for c, _ in results[:10]).lower()

print("="*80)
print(f"ID: {mid}")
print(f"Q: {question}")
for letter, text in sorted(options.items()):
    opt_lower = text.lower()
    found = opt_lower in evidence
    key_words = [w for w in re.findall(r'\w+', opt_lower) if len(w) > 2]
    matches = sum(1 for w in key_words if w in evidence)
    print(f"  {letter}: {text}")
    print(f"     exact_match={found}, keyword_matches={matches}/{len(key_words)}")

print()
print("TOP 3 EVIDENCE:")
for i, (chunk, score) in enumerate(results[:3]):
    print(f"  [{i}] score={score:.2f} doc={chunk.doc_id}")
    print(f"      {chunk.text[:300]}")
    print()

# Check id=736 (learning rate quá nhỏ)
mid = 736
q_data = questions[float(mid)]
question = q_data["fun_question"]
note = q_data.get("note", "") or ""
options = parse_options(note)

results = retriever.retrieve_with_metadata_filter(query=question, top_k=10)
print("="*80)
print(f"ID: {mid}")
print(f"Q: {question}")
print("TOP 3 EVIDENCE:")
for i, (chunk, score) in enumerate(results[:3]):
    print(f"  [{i}] score={score:.2f} doc={chunk.doc_id}")
    print(f"      {chunk.text[:400]}")
    print()
