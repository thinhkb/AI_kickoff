"""Debug negation and regression cases."""
import sys, json, re
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

from src.document.retriever import DocumentRetriever
from src.document.option_parser import parse_options
from src.document.option_scorer import OptionScorer

retriever = DocumentRetriever()
retriever.load_index(PROJECT_ROOT / "data" / "processed" / "document_chunks.jsonl")
scorer = OptionScorer()

import openpyxl
wb = openpyxl.load_workbook(PROJECT_ROOT / "data" / "example_data" / "example_data.xlsx")
ws = wb["example_question"]
headers = [c.value for c in ws[1]]
questions = {}
for row in ws.iter_rows(min_row=2, values_only=True):
    d = dict(zip(headers, row))
    questions[d["id"]] = d

# Check 713 (new regression)
for mid in [713, 715, 733, 742]:
    mid_float = float(mid)
    if mid_float not in questions:
        continue
    q_data = questions[mid_float]
    question = q_data["fun_question"]
    note = q_data.get("note", "") or ""
    
    options = parse_options(note)
    is_neg = OptionScorer.is_negation_question(question)
    
    # Extract doc_id from note
    doc_id = retriever._detect_doc_id(question) or retriever._detect_doc_id(note)
    
    results = retriever.retrieve_with_metadata_filter(query=question, doc_id=doc_id, top_k=50)
    chunks = [chunk for chunk, score in results]
    
    scores = scorer.score_options(
        question=question, options=options,
        evidence_chunks=chunks, retriever=retriever,
    )
    
    sorted_scores = sorted(scores.items(), key=lambda x: x[1])
    selected = sorted_scores[0][0] if is_neg else sorted(scores.items(), key=lambda x: -x[1])[0][0]
    
    print(f"ID: {mid} | neg={is_neg} | selected={selected}")
    print(f"Q: {question[:120]}")
    for letter in sorted(options.keys()):
        print(f"  {letter}: {options[letter][:80]} | score={scores.get(letter, 0):.4f}")
    print()
