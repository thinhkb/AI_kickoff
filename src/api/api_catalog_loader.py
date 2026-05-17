"""
Loads API definitions from Excel and converts them into a structured registry.
"""
import json
import re
from pathlib import Path
from typing import List, Dict, Any

from src.schemas import APIEntry
from src.utils.io_utils import read_excel_sheet
from src.utils.logging_utils import logger
from configs.paths import API_CONFIG_FILE


def load_api_catalog(
    path: str | Path = None,
) -> List[APIEntry]:
    """
    Load API definitions from the Excel config file.
    Returns a list of APIEntry objects.
    """
    path = Path(path or API_CONFIG_FILE)
    logger.info(f"Loading API catalog from {path}...")

    rows = read_excel_sheet(path, sheet_name="Doc_api_for_contest")
    logger.info(f"  Found {len(rows)} API entries")

    apis = []
    for row in rows:
        if not row.get("func_code"):
            continue
        try:
            api = APIEntry.from_dict(row)
            apis.append(api)
        except Exception as e:
            logger.warning(f"  Skipping invalid API entry: {e}")

    logger.info(f"  Loaded {len(apis)} valid APIs")
    return apis


def load_alias_dictionary(
    path: str | Path = None,
) -> Dict[str, Dict[str, str]]:
    """
    Load alias dictionaries from the 'Doc_alias_for_contest' sheet.

    Returns a dict of {param_name: {display_key: backend_value}}.
    Example:
    {
        "organization": {"Trung tâm phần mềm viễn thông": "TTPMVT", ...},
        "projectType": {"T&M": "T&M", "Presales": "presales", ...},
        ...
    }
    """
    path = Path(path or API_CONFIG_FILE)
    logger.info(f"Loading alias dictionary from {path}...")

    import openpyxl
    wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
    ws = wb["Doc_alias_for_contest"]

    alias_dict = {}

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column, values_only=True):
        for cell_value in row:
            if cell_value is None:
                continue
            cell_str = str(cell_value).strip()
            if not cell_str:
                continue

            # Parse entries like: paramName = [\n {key: ..., value: ...}, ...\n]
            parsed = _parse_alias_entry(cell_str)
            if parsed:
                name, mapping = parsed
                alias_dict[name] = mapping

    wb.close()
    logger.info(f"  Loaded {len(alias_dict)} alias categories")
    return alias_dict


def _parse_alias_entry(text: str) -> tuple | None:
    """
    Parse a single alias entry like:
    'organization = [\n    {"key": "Trung tâm...", "value": "TTPMVT"}, ...\n]'

    Returns (param_name, {display_key: backend_value}) or None.
    """
    # Match paramName = [...]
    m = re.match(r"(\w+)\s*=\s*\[", text)
    if not m:
        return None

    param_name = m.group(1)

    # Extract the JSON array part
    bracket_start = text.index("[")
    # Find matching ]
    bracket_count = 0
    bracket_end = -1
    for i in range(bracket_start, len(text)):
        if text[i] == "[":
            bracket_count += 1
        elif text[i] == "]":
            bracket_count -= 1
            if bracket_count == 0:
                bracket_end = i
                break

    if bracket_end == -1:
        # Try adding a closing bracket
        array_str = text[bracket_start:] + "]"
    else:
        array_str = text[bracket_start:bracket_end + 1]

    # Clean up Python-style dicts to JSON
    array_str = array_str.replace("'", '"')
    # Handle trailing commas
    array_str = re.sub(r",\s*([}\]])", r"\1", array_str)

    try:
        items = json.loads(array_str)
    except json.JSONDecodeError:
        # Try to extract key-value pairs with regex
        items = []
        for match in re.finditer(
            r'"key"\s*:\s*"([^"]*?)"\s*,\s*"value"\s*:\s*"([^"]*?)"',
            text,
        ):
            items.append({"key": match.group(1), "value": match.group(2)})

    if not items:
        return None

    mapping = {}
    for item in items:
        if isinstance(item, dict):
            key = item.get("key", "")
            value = item.get("value", "")
            if key or value:
                mapping[key] = value
            # Also handle projectName/projectId format
            if "projectName" in item:
                mapping[item["projectName"]] = str(item.get("projectId", ""))

    return param_name, mapping


def build_api_search_text(api: APIEntry) -> str:
    """
    Build a combined text representation for API retrieval.
    """
    parts = [
        api.func_code,
        api.name,
        api.description,
        api.example_question,
    ]
    return " ".join(p for p in parts if p)
